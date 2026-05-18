#!/usr/bin/env python3
"""
importar-facturas.py — Importa facturas desde Movimientos_caja.xlsx a la BD
y copia los PDFs a uploads/facturas/.

Idempotente: omite filas cuyo nombre_archivo ya existe en BD y filas sin PDF.
Ejecuta un único batch SQL para mayor eficiencia.

Uso:
    python3 /volume1/docker/club/new/scripts/importar-facturas.py
"""

import os
import re
import shutil
import subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    subprocess.run(['pip3', 'install', 'openpyxl', '-q'], check=True)
    import openpyxl

SCRIPTS_DIR  = '/volume1/docker/club/new/scripts'
UPLOADS_DIR  = '/volume1/docker/club/new/uploads/facturas'
EXCEL_FILE   = os.path.join(SCRIPTS_DIR, 'Movimientos_caja.xlsx')
DB_CONTAINER = 'club-db-1'
DB_USER      = 'deporte'
DB_NAME      = 'deporte_pedrola'
APP_UPLOADS  = '/app/uploads/facturas'   # ruta dentro del contenedor app


def psql_batch(sql_text):
    """Ejecuta un bloque SQL en el contenedor PostgreSQL de una sola vez."""
    result = subprocess.run(
        ['docker', 'exec', '-i', DB_CONTAINER,
         'psql', '-U', DB_USER, '-d', DB_NAME, '--no-psqlrc'],
        input=sql_text, capture_output=True, text=True, timeout=120
    )
    return result.stdout, result.stderr, result.returncode


def q(v):
    """Escapa un valor para SQL: devuelve 'texto' o NULL."""
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def num(v):
    """Convierte a literal numérico SQL o NULL."""
    if v is None:
        return 'NULL'
    try:
        return f'{float(v):.2f}'
    except (TypeError, ValueError):
        return 'NULL'


def get_filename(cell):
    """Extrae el nombre de archivo de una celda (texto plano o fórmula HYPERLINK)."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str) and v.startswith('=HYPERLINK'):
        m = re.search(r',"([^"]+)"\)', v)
        return m.group(1) if m else None
    return str(v).strip()


def main():
    os.makedirs(UPLOADS_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Movimientos']
    total_filas = ws.max_row - 1

    print(f'\n=== Importando facturas desde Excel ({total_filas} filas) ===')
    print(f'    Origen PDFs : {SCRIPTS_DIR}')
    print(f'    Destino PDFs: {UPLOADS_DIR}\n')

    # ── Paso 1: obtener los nombre_archivo ya en BD ───────────────────────────
    out, err, code = psql_batch('SELECT nombre_archivo FROM facturas;')
    if code != 0:
        print(f'ERROR conectando a la BD: {err}')
        return
    ya_en_bd = set(
        line.strip() for line in out.splitlines()
        if line.strip()
        and 'nombre_archivo' not in line
        and line.strip() != '---'
        and 'row' not in line
        and '(' not in line
    )
    print(f'    Facturas ya en BD: {len(ya_en_bd)}')

    # ── Paso 2: procesar filas del Excel ─────────────────────────────────────
    inserts       = []
    pdfs_a_copiar = []
    ya_existia    = 0
    sin_pdf       = 0
    sin_pdf_list  = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        fname = get_filename(row[0])
        if not fname:
            continue

        tipo       = row[1].value          # factura / recibo
        concepto   = row[2].value
        proveedor  = row[3].value
        base       = row[4].value          # Importe (€) = base imponible
        iva_import = row[5].value          # IVA (€)
        # columna 6 = fórmula =E+F, la recalculamos nosotros
        deporte    = row[7].value
        equipo_cat = row[8].value
        fecha      = row[9].value

        base_f = float(base)       if base       is not None else None
        iva_f  = float(iva_import) if iva_import is not None else None

        if base_f is not None and iva_f is not None:
            importe_total = base_f + iva_f
            iva_pct = round(iva_f / base_f * 100, 2) if base_f > 0 and iva_f > 0 else None
        else:
            importe_total = base_f
            iva_pct = None

        src_path = os.path.join(SCRIPTS_DIR, fname)
        if not os.path.exists(src_path):
            sin_pdf += 1
            sin_pdf_list.append(fname)
            continue

        if fname in ya_en_bd:
            ya_existia += 1
            continue

        fecha_sql = q(fecha.strftime('%Y-%m-%d')) if isinstance(fecha, datetime) else 'NULL'
        app_path  = f'{APP_UPLOADS}/{fname}'

        sql = (
            "INSERT INTO facturas "
            "  (nombre_archivo, ruta_archivo, tipo, proveedor, concepto, "
            "   base_imponible, iva_porcentaje, iva_importe, importe, "
            "   deporte, equipo_categoria, fecha_factura, ocr_revisado) "
            "VALUES ("
            f"  {q(fname)}, {q(app_path)}, {q(tipo)}, {q(proveedor)}, {q(concepto)}, "
            f"  {num(base_f)}, {num(iva_pct)}, {num(iva_f)}, {num(importe_total)}, "
            f"  {q(deporte)}, {q(equipo_cat)}, {fecha_sql}, true"
            ") ON CONFLICT DO NOTHING;"
        )
        inserts.append((fname, sql))
        pdfs_a_copiar.append((src_path, os.path.join(UPLOADS_DIR, fname)))

    print(f'    A importar : {len(inserts)}')
    print(f'    Ya existían: {ya_existia}')
    print(f'    Sin PDF    : {sin_pdf}  (pendientes de subir)\n')

    if not inserts:
        print('Nada nuevo que importar.')
        if sin_pdf:
            print(f'Sube los {sin_pdf} PDFs restantes y vuelve a ejecutar.')
        return

    # ── Paso 3: copiar PDFs ───────────────────────────────────────────────────
    print('Copiando PDFs...')
    for src, dst in pdfs_a_copiar:
        shutil.copy2(src, dst)
    print(f'  {len(pdfs_a_copiar)} PDFs copiados')

    # ── Paso 4: insertar en BD (un único batch con transacción) ──────────────
    print('Insertando en base de datos...')
    sql_batch = 'BEGIN;\n' + '\n'.join(s for _, s in inserts) + '\nCOMMIT;\n'
    out, err, code = psql_batch(sql_batch)

    if code == 0:
        importadas = len(inserts)
        print(f'  OK: {importadas} registros insertados')
    else:
        print(f'  ERROR en la inserción:\n{err}')
        # Revertir los PDFs copiados
        for _, dst in pdfs_a_copiar:
            if os.path.exists(dst):
                os.remove(dst)
        importadas = 0

    print()
    print('=' * 52)
    print(f'  Importadas   : {importadas}')
    print(f'  Ya existían  : {ya_existia}')
    print(f'  Sin PDF      : {sin_pdf}  <- pendientes de subir')
    print(f'  Total Excel  : {total_filas}')
    print('=' * 52)

    if sin_pdf_list:
        print(f'\nPDFs pendientes (primeros 30):')
        for n in sin_pdf_list[:30]:
            print(f'  - {n}')
        if len(sin_pdf_list) > 30:
            print(f'  ... y {len(sin_pdf_list) - 30} más')
        print('\nVuelve a ejecutar el script cuando hayas subido más PDFs.')
    print()


if __name__ == '__main__':
    main()
